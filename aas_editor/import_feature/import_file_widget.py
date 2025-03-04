#  Copyright (C) 2022  Igor Garmaev, garmaev@gmx.net
#
#  This program is made available under the terms of the GNU General Public License as published by
#  the Free Software Foundation, either version 3 of the License, or (at your option) any later version.
#
#  This program is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
#  without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
#
#  A copy of the GNU General Public License is available at http://www.gnu.org/licenses/
from dataclasses import dataclass

import openpyxl
from PyQt6.QtCore import QModelIndex
from PyQt6.QtGui import QIntValidator
from PyQt6.QtWidgets import QWidget, QHBoxLayout, QPushButton, QFileDialog, QMessageBox, QFormLayout, QDialog, \
    QDialogButtonBox, QLabel, QLineEdit

from aas_editor.import_feature import import_util
from aas_editor.import_feature.import_settings import MAPPING_ATTR
from aas_editor.import_feature import import_util_classes
from aas_editor.import_feature.import_util import getMapping, usedColumnsInMapping, unusedColumnsInMapping
from aas_editor.package import Package
from aas_editor import settings
from aas_editor.utils.util_classes import ClassesInfo
from aas_editor.utils.util_type import isIterable
from aas_editor import dialogs

CHOOSE = "Choose..."
EXCEL_FILES = "Excel files (*.xlsx)"


@dataclass
class ImportSettings:
    def __init__(self,
                 mappingPackage: Package = None,
                 mappingFile: str = None,
                 sourceFile: str = None,
                 sheetname: str = None,  # required if sourcefile is Excel
                 exampleRow: int = 2):
        self.mappingPackage = mappingPackage
        self.mappingFile = mappingFile
        self.sourceFile = sourceFile
        self.sheetname = sheetname
        self.exampleRow = exampleRow

    @property
    def exampleRow(self):
        return self._exampleRow

    @exampleRow.setter
    def exampleRow(self, row):
        self._exampleRow = row
        if self.sourceFile:
            self.exampleRowValue = import_util.importRowValueFromExcel(sourcefile=self.sourceFile, row=row)
            import_util_classes.PreObjectImport.EXAMPLE_ROW_VALUE = self.exampleRowValue

    @property
    def sheetname(self):
        return self._sheetname

    @sheetname.setter
    def sheetname(self, name):
        if self.sourceFile:
            excelFile = openpyxl.load_workbook(self.sourceFile, data_only=True)
            if name:
                if name not in excelFile.sheetnames:
                    raise ValueError(f"There is no such sheetname in the sourcefile: {name}")
                self._sheetname = name
            else:
                self._sheetname = excelFile.sheetnames[0]
        else:
            self._sheetname = name


class ImportManageWidget(QWidget):
    IMPORT_SETTINGS = ImportSettings()

    def __init__(self, parent: "ImportApp" = None, mainTreeView=None):
        super(ImportManageWidget, self).__init__(parent)
        self.importApp = parent
        self.setupLayout()

        self.importFileLine = QLineEdit(self, placeholderText="Choose the file to import from")
        self.importFileLine.setReadOnly(True)
        self.saveMappingBtn = QPushButton("Save mapping", self, clicked=self.saveMappingFileDialog)
        self.importSettingsBtn = QPushButton("Settings", self, clicked=self.initSettingsDialog)
        self.importBtn = QPushButton("Run Import", self, clicked=self.initRunImportDialog)
        for widget in (self.importFileLine, self.saveMappingBtn, self.importBtn, self.importSettingsBtn):
            self.layout().addWidget(widget)

        self.mainTreeView: "ImportTreeView" = mainTreeView
        self.importFile: str = ""
        self.fileType: str = ""

    def setTreeView(self, treeview):
        self.mainTreeView = treeview

    def setupLayout(self):
        toolBarHLayout = QHBoxLayout()
        toolBarHLayout.setContentsMargins(5, 0, 5, 0)
        self.setFixedHeight(settings.TOOLBARS_HEIGHT)
        self.setLayout(toolBarHLayout)

    def saveMappingFileDialog(self):
        file, _ = QFileDialog.getSaveFileName(self, 'Save Mapping File', filter=settings.JSON_FILES_FILTER)
        if file:
            self.saveMappingFile(file)

    def saveMappingFile(self, file):
        try:
            import_util.saveMapping(pack=ImportManageWidget.IMPORT_SETTINGS.mappingPackage, file=file)
        except (TypeError, ValueError, KeyError) as e:
            dialogs.ErrorMessageBox.withTraceback(self, f"Package couldn't be saved: {file}: {e}").exec()
        except AttributeError as e:
            dialogs.ErrorMessageBox.withTraceback(self, f"No chosen package to save: {e}").exec()

    def execImportSettingsDialog(self) -> bool:
        dialog = ImportSettingsDialog(self)
        result = False
        while not result and dialog.exec() == QDialog.DialogCode.Accepted:
            try:
                try:
                    aasFile = dialog.aasFileLine.text()
                    self.importApp.mainTreeView.closeAllFiles()
                    ImportManageWidget.IMPORT_SETTINGS.mappingPackage = self.importApp.mainTreeView.openPack(aasFile)
                    self.importApp.packTreeModel.setData(QModelIndex(), [], settings.UNDO_ROLE)
                except Exception as e:
                    dialogs.ErrorMessageBox.withTraceback(self, f"Could not open AAS File: {e}").exec()
                    continue

                ImportManageWidget.IMPORT_SETTINGS.sourceFile = dialog.importExcelFileLine.text()
                try:
                    ImportManageWidget.IMPORT_SETTINGS.sheetname = dialog.sheetnameLine.text()
                except Exception as e:
                    QMessageBox.critical(self, "Error: wrong sheetname", str(e))
                    continue
                ImportManageWidget.IMPORT_SETTINGS.exampleRow = int(dialog.exampleRow.text())

                try:
                    mappingFile = dialog.mappingFileLine.text()
                    if mappingFile:
                        import_util.setMappingFromFile(pack=ImportManageWidget.IMPORT_SETTINGS.mappingPackage,
                                                       mappingFile=mappingFile)
                        ImportManageWidget.IMPORT_SETTINGS.mappingFile = mappingFile
                except Exception as e:
                    dialogs.ErrorMessageBox.withTraceback(self, f"Could not open Mapping File: {e}").exec()
                    continue

                result = True
            except Exception as e:
                dialogs.ErrorMessageBox.withTraceback(self, str(e)).exec()
                continue
        return result

    def initSettingsDialog(self):
        dialog = SettingsDialog(self, importSettings=ImportManageWidget.IMPORT_SETTINGS)
        result = False
        while not result and dialog.exec() == QDialog.DialogCode.Accepted:
            try:
                ImportManageWidget.IMPORT_SETTINGS.sourceFile = dialog.importExcelFileLine.text()
                ImportManageWidget.IMPORT_SETTINGS.sheetname = dialog.sheetnameLine.text()
                ImportManageWidget.IMPORT_SETTINGS.exampleRow = int(dialog.exampleRow.text())
                try:
                    mappingFile = dialog.mappingFileLine.text()
                    if mappingFile and mappingFile != ImportManageWidget.IMPORT_SETTINGS.mappingFile:
                        import_util.setMappingFromFile(pack=ImportManageWidget.IMPORT_SETTINGS.mappingPackage,
                                                       mappingFile=mappingFile)
                except Exception as e:
                    dialogs.ErrorMessageBox.withTraceback(self, f"Could not open Mapping File: {e}").exec()
                    continue

                result = 1
            except Exception as e:
                dialogs.ErrorMessageBox.withTraceback(self, str(e)).exec()
                continue

    def importFromPack(self, templatePack: Package, sourcefile: str, minRow: int, maxRow: int,
                       exportFolder: str, fileNameScheme: str):
        excelfile = openpyxl.load_workbook(sourcefile, data_only=True)
        if maxRow == -1:
            sheet = excelfile[ImportManageWidget.IMPORT_SETTINGS.sheetname]
            maxRow = sheet.max_row
        for row in range(minRow, maxRow + 1):
            try:
                newPack = self.newPackFromTemplate(templatePack, row, sourceWB=excelfile)
                if not fileNameScheme:
                    fileName = f"{row}.aasx"
                else:
                    fileName = import_util.importValueFromExcelWB(
                        fileNameScheme, workbook=excelfile, row=row,
                        sheetname=ImportManageWidget.IMPORT_SETTINGS.sheetname)
                    if "." not in fileName:
                        fileName = f"{fileName}.aasx"
                newPack.write(f"{exportFolder}/{fileName}")
            except Exception as e:
                e.add_note(f"Problem occurred by importing values from the row {row}")
                raise e

        QMessageBox.information(self, "Export was successful", f"{maxRow-minRow+1} AAS files were successfully generated!")

    def newPackFromTemplate(self, pack: Package, row: int, sourceWB: openpyxl.Workbook):
        newPack = Package()
        for refObj in pack.objStore:
            newRefObj = self._initObjWithMappingImport(refObj, row, sourceWB=sourceWB)
            newPack.objStore.add(newRefObj)
        return newPack

    def _initObjWithMappingImport(self, obj, row: int, sourceWB: openpyxl.Workbook):
        preobj = import_util_classes.PreObjectImport.fromObject(obj, withIterParams=False)
        preobj.setMapping(getattr(obj, MAPPING_ATTR, {}))
        newObj = preobj.initWithImport(rowNum=row, sourceWB=sourceWB,
                                       sheetname=ImportManageWidget.IMPORT_SETTINGS.sheetname)

        if isIterable(obj) and ClassesInfo.iterAttrs(type(obj)):
            iterAttrs = ClassesInfo.iterAttrs(type(obj))  # TODO if pyi40aas changes
            for iterAttr in iterAttrs:
                for iterObj in getattr(obj, iterAttr):
                    newIterObj = self._initObjWithMappingImport(iterObj, row=row, sourceWB=sourceWB)
                    if hasattr(newIterObj, "id_short") and isinstance(newIterObj.id_short, str) and (
                            newIterObj.id_short.startswith("generated_submodel_list_hack_")):  # FIXME: Refactor after basyx fix
                        newIterObj.id_short = None
                    getattr(newObj, iterAttr).add(newIterObj)
        return newObj

    def initRunImportDialog(self):
        dialog = RunImportDialog(self, importSettings=ImportManageWidget.IMPORT_SETTINGS)
        result = False
        while not result and dialog.exec() == QDialog.DialogCode.Accepted:
            try:
                self.importFromPack(templatePack=ImportManageWidget.IMPORT_SETTINGS.mappingPackage,
                                    sourcefile=ImportManageWidget.IMPORT_SETTINGS.sourceFile,
                                    minRow=int(dialog.minRow.text()),
                                    maxRow=int(dialog.maxRow.text()),
                                    exportFolder=dialog.exportFolderLine.text(),
                                    fileNameScheme=dialog.nameScheme.text()
                                    )
            except Exception as e:
                dialogs.ErrorMessageBox.withTraceback(self, str(e)).exec()
                continue


class SettingsDialog(QDialog):
    def __init__(self, parent=None, title="AAS Manager Import Tool", *, importSettings: ImportSettings):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.buttonBox = QDialogButtonBox(QDialogButtonBox.StandardButton.Apply | QDialogButtonBox.StandardButton.Cancel, self)
        self.applyBtn = self.buttonBox.button(QDialogButtonBox.StandardButton.Apply)
        self.cancelBtn = self.buttonBox.button(QDialogButtonBox.StandardButton.Cancel)
        self.applyBtn.released.connect(self.accept)
        self.cancelBtn.released.connect(self.reject)

        # 6x4
        self.formLayout = QFormLayout()
        # Create widgets
        self.importSourceFileGB(importSettings.sourceFile)
        self.sheetnameGB(importSettings.sheetname)
        self.exampleRowGB(importSettings.exampleRow)
        self.mappingFileGB(importSettings.mappingFile)

        self.formLayout.addWidget(self.buttonBox)
        self.setLayout(self.formLayout)

    def importSourceFileGB(self, file=None):
        self.importExcelFileLine = QLineEdit(placeholderText="Excel file")
        self.importExcelFileLine.setText(file)
        self.chooseSourceFileBtn = QPushButton(CHOOSE, clicked=self.chooseImportFile)
        hbox = QHBoxLayout()
        hbox.addWidget(self.importExcelFileLine)
        hbox.addWidget(self.chooseSourceFileBtn)
        self.formLayout.addRow(QLabel("Excel to import values from*"), hbox)

    def exampleRowGB(self, exampleRow=None):
        self.exampleRow = QLineEdit(self,
                                    placeholderText="Row for example values (normally=2)")
        self.exampleRow.setValidator(QIntValidator(1, 100000))
        if exampleRow:
            self.exampleRow.setText(str(exampleRow))
        self.formLayout.addRow("Example Row*", self.exampleRow)

    def sheetnameGB(self, sheetname=None):
        self.sheetnameLine = QLineEdit(placeholderText="Leave empty if first Excel sheet used")
        self.sheetnameLine.setText(sheetname)
        self.formLayout.addRow("Sheetname", self.sheetnameLine)

    def mappingFileGB(self, file=None):
        self.mappingFileLine = QLineEdit(placeholderText="Optional")
        self.mappingFileLine.setText(file)
        self.chooseImportFileBtn = QPushButton(CHOOSE, clicked=self.chooseMappingFile)

        hbox = QHBoxLayout()
        hbox.addWidget(self.mappingFileLine)
        hbox.addWidget(self.chooseImportFileBtn)
        self.formLayout.addRow("Mapping File", hbox)

    def chooseImportFile(self):
        file, _ = QFileDialog.getOpenFileName(self, "Choose Import Excel File",
                                              filter=f"{EXCEL_FILES};;{settings.ALL_FILES_FILTER}")
        if file:
            self.importExcelFileLine.setText(file)

    def chooseMappingFile(self):
        file, _ = QFileDialog.getOpenFileName(self, "Choose Mapping File", filter=settings.JSON_FILES_FILTER)
        if file:
            self.mappingFileLine.setText(file)


class ImportSettingsDialog(SettingsDialog):
    def __init__(self, parent=None, title="AAS Manager Excel Import Tool"):
        super(SettingsDialog, self).__init__(parent)
        self.setWindowTitle(title)
        self.buttonBox = QDialogButtonBox(QDialogButtonBox.StandardButton.Apply, self)
        self.applyBtn = self.buttonBox.button(QDialogButtonBox.StandardButton.Apply)
        self.applyBtn.released.connect(self.accept)

        # 6x4
        self.formLayout = QFormLayout()
        self.formLayout.addRow(QLabel("The tool provides possibility to generate AAS files for each row in Excel table:\n"
                                      "It takes Base AAS File as a Template with static information. \n"
                                      "You can edit the Base AAS as a normal AAS and create there new Submodels, Properties etc.. \n"
                                      "You can input there static values, which will be the same in all generated AAS \n"
                                      "or you can input there Excel column references for specific attributes of objects.\n"
                                      "When you run generation, for each row in Excel table an AAS based on Base AAS will be created. \n"
                                      "The values in generated AAS will be token from Base AAS. \n"
                                      "If Base AAS has parameters with Excel-Column References instead of static values, \n"
                                      "then in generated AAS for each row these parameters will be set to values from corresponding column of Excel-Table. \n\n"
                                      "E.g. you can set value of property to $A$, then each generated AAS will have Property with value from column A and corresponding row. \n"
                                      "You can also mix static values und multiple references. E.g. set value of property to $A$_$B$_example "))
        # Create widgets
        self.aasFileGB()
        self.importSourceFileGB()
        self.sheetnameGB()
        self.exampleRowGB()
        self.mappingFileGB()

        self.formLayout.addWidget(self.buttonBox)
        self.setLayout(self.formLayout)

    def aasFileGB(self):
        self.aasFileLine = QLineEdit(placeholderText="Use existing or create new AAS file")

        self.chooseAasFileBtn = QPushButton(CHOOSE, clicked=self.chooseAasFile)
        self.newAasFileBtn = QPushButton("New...", clicked=self.newAasFile)

        hbox = QHBoxLayout()
        hbox.addWidget(self.aasFileLine)
        hbox.addWidget(self.chooseAasFileBtn)
        hbox.addWidget(self.newAasFileBtn)
        self.formLayout.addRow("Base AAS File*", hbox)

    def newAasFile(self):
        file, _ = QFileDialog.getSaveFileName(self, 'Create new AAS File', 'new_aas_file.aasx',
                                              filter=settings.FILTER_AAS_FILES)
        if file:
            Package().write(file)
            self.aasFileLine.setText(file)
        else:  # cancel pressed
            return

    def chooseAasFile(self):
        file, _ = QFileDialog.getOpenFileName(self, "Choose Import Excel File", filter=settings.FILTER_AAS_FILES)
        if file:
            self.aasFileLine.setText(file)


class RunImportDialog(QDialog):
    def __init__(self, parent: ImportManageWidget, *, importSettings: ImportSettings):
        super().__init__(parent)
        self.setWindowTitle("Run import")
        self.buttonBox = QDialogButtonBox(QDialogButtonBox.StandardButton.Cancel | QDialogButtonBox.StandardButton.Apply, self)
        self.applyBtn = self.buttonBox.button(QDialogButtonBox.StandardButton.Apply)
        self.cancelBtn = self.buttonBox.button(QDialogButtonBox.StandardButton.Cancel)
        self.applyBtn.released.connect(self.accept)
        self.cancelBtn.released.connect(self.reject)

        layout = QFormLayout(self)

        self.minRow = QLineEdit(self, placeholderText="Row where value import starts")
        self.minRow.setValidator(QIntValidator(1, 100000))
        layout.addRow("Start Row*", self.minRow)

        self.maxRow = QLineEdit(self, placeholderText="Row where value import ends (-1 for the last row)")
        self.maxRow.setValidator(QIntValidator(-1, 100001))
        layout.addRow("End Row*", self.maxRow)

        self.exportFolderLine = QLineEdit(placeholderText="Export Folder")
        hbox = QHBoxLayout()
        hbox.addWidget(self.exportFolderLine)
        self.chooseExportFolderBtn = QPushButton("Choose Export Folder", clicked=self.chooseExportFolder)
        hbox.addWidget(self.chooseExportFolderBtn)
        layout.addRow("Export Folder*", hbox)

        self.nameScheme = QLineEdit(self, placeholderText="e.g. AAS_file_$A$.aasx")
        layout.addRow("Generated AAS filenames scheme", self.nameScheme)

        packMapping = getMapping(importSettings.mappingPackage)
        self.usedColsLine = QLabel(str(usedColumnsInMapping(packMapping)).strip("[]").replace("'", ""), self)
        layout.addRow("Used columns for value import", self.usedColsLine)

        self.unusedColsLine = QLabel(str(unusedColumnsInMapping(packMapping, sourcefile=importSettings.sourceFile,
                                     sheetname=importSettings.sheetname)).strip("[]").replace("'", ""), self)
        layout.addRow("Unused columns for value import", self.unusedColsLine)

        layout.addWidget(self.buttonBox)
        self.setLayout(layout)

    def chooseExportFolder(self):
        folder = QFileDialog.getExistingDirectory(self, "Choose Export Folder")
        if folder:
            self.exportFolderLine.setText(folder)
