#  Copyright (C) 2022  Igor Garmaev, garmaev@gmx.net
#
#  This program is made available under the terms of the GNU General Public License as published by
#  the Free Software Foundation, either version 3 of the License, or (at your option) any later version.
#
#  This program is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
#  without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
#
#  A copy of the GNU General Public License is available at http://www.gnu.org/licenses/
#
#  This program is made available under the terms of the GNU General Public License as published by
#  the Free Software Foundation, either version 3 of the License, or (at your option) any later version.
#
#  This program is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY;
#  without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.
#
#  A copy of the GNU General Public License is available at http://www.gnu.org/licenses/
import json
import re
from typing import List, Dict

import basyx
import openpyxl
from basyx.aas.model import ModelReference, Key, KeyTypes
from openpyxl.worksheet.worksheet import Worksheet

from aas_editor.tools.import_feature import import_settings
from aas_editor.utils import util_type

COLUMNS_PATTERN = re.compile(r"\$[A-Z][A-Z]?\$")


def colLettersInExcelSheet(sheet: Worksheet) -> List[str]:
    colLetters = []
    for columnNum in range(1, sheet.max_column + 1):
        colLetters.append(openpyxl.utils.cell.get_column_letter(columnNum))
    return colLetters


def importRowValueFromExcel(sourcefile, sheetname=None, row=2):
    excel_file = openpyxl.load_workbook(sourcefile, data_only=True)
    if not sheetname:
        sheetname = excel_file.sheetnames[0]
    sheet = excel_file[sheetname]

    rowValue: Dict = {}
    for colLetter in colLettersInExcelSheet(sheet):
        rowValue[colLetter] = sheet[f"{colLetter}{row}"].value
    return rowValue


def _mapping4referableIntoDict(obj, mapDict):
    mapping = getattr(obj, import_settings.MAPPING_ATTR, {})
    if mapping:
        ref = ModelReference.from_referable(obj) # FIXME V30RC02
        keys = ','.join([f"Key(type_={i.type}, value='{i.value}')" for i in ref.key])
        target_type = repr(ref.type).lstrip("<class '").rstrip("'>")
        mapDict[f"ModelReference(type_={target_type}, key=({keys},))"] = mapping


def _mappingIntoDict(obj, mapDict):
    """Save all existing mappings in the obj into mapDict"""
    _mapping4referableIntoDict(obj, mapDict)
    if util_type.isIterable(obj):
        for i in obj:
            _mappingIntoDict(i, mapDict)


def getMapping(pack: "Package"):
    mapDict = dict()
    for obj in pack.objStore:
        _mappingIntoDict(obj, mapDict)
    return mapDict


def saveMapping(pack: "Package", file: str) -> bool:
    mapDict = getMapping(pack)
    with open(file, 'w') as jsonFile:
        json.dump(mapDict, jsonFile)
        jsonFile.close()
    return True


def usedColumnsInMapping(mapDict) -> List[str]:
    mapping = str(mapDict)
    columns: List[str] = re.findall(COLUMNS_PATTERN, mapping)
    columns = list(set([col.strip("$") for col in columns]))
    columns.sort()
    columns.sort(key=len)
    return columns


def unusedColumnsInMapping(mapDict, sourcefile, sheetname: Worksheet) -> List[str]:
    excel_file = openpyxl.load_workbook(sourcefile, data_only=True)
    if not sheetname:
        sheetname = excel_file.sheetnames[0]
    sheet = excel_file[sheetname]

    usedColumns = usedColumnsInMapping(mapDict)
    columns = colLettersInExcelSheet(sheet)
    for col in usedColumns:
        columns.remove(col)
    columns.sort(key=len)
    return columns


def setMappingFromFile(pack: "Package", mappingFile: str):
    with open(mappingFile, 'r') as jsonFile:
        mapDict = json.load(jsonFile)

    for refRepr in mapDict:
        aasref: ModelReference = eval(refRepr, {
            "Key": Key,
            "KeyTypes": KeyTypes,
            "ModelReference": ModelReference,
            "basyx": basyx,
        })
        refObj = aasref.resolve(pack.objStore)
        mapping = mapDict[refRepr]
        setattr(refObj, import_settings.MAPPING_ATTR, mapping)


def importValueFromExampleRow(value: str, row: Dict):
    """
    Import value from a row based on a given raw value string.

    e.g. if rawValue == "$A$" -> return row["A"]
    if rawValue == "$A$ some text $B$" -> return f"{str(row["A"])} some text {str(row["B"])}"

    :param raw_value: The raw value string containing column references.
    :param row: The dictionary representing the row with column data.
    :return: The imported value with column references replaced by actual values.
    """
    colReferences: List[str] = re.findall(COLUMNS_PATTERN, value)
    for col in colReferences:
        column = col.strip("$")
        importedVal = str(row[column])
        value = value.replace(f"${column}$", importedVal, -1)
    return value


def importValueFromExcelWB(value: str, workbook: openpyxl.Workbook, sheetname, row=2):
    """
    Import value from an Excel workbook based on a given value string.

    :param value: The value string containing column references.
    :param workbook: The openpyxl Workbook object.
    :param sheetname: The name of the sheet from which to import values.
    :param row: The row number to import values from (default is 2).
    :return: The imported value with column references replaced by actual values.
    """
    sheet = workbook[sheetname]
    colReferences: List[str] = re.findall(COLUMNS_PATTERN, value)
    for col in colReferences:
        column = col.strip("$")
        importedVal = str(sheet[f"{column}{row}"].value)
        value = value.replace(f"${column}$", importedVal, -1)
    return value


def importValueFromExcel(value: str, sourcefile, sheetname, row=2):
    excel_file = openpyxl.load_workbook(sourcefile, data_only=True)
    return importValueFromExcelWB(value, excel_file, sheetname, row)


def isValueToImport(value):
    if re.search(COLUMNS_PATTERN, value):
        return True
    return False
